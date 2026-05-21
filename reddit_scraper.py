import argparse
import random
import time
from datetime import date, datetime, timedelta

import httpx
import openpyxl
from langdetect import detect, LangDetectException

XLSX_PATH = "Data gathering example(1).xlsx"
REDDIT_SHEET = "Reddit"
POSTS_PER_KEYWORD = 40
DATA_START_ROW = 7  # rows 1-6 are account info + blank + headers

KEYWORDS = [
    "Economy",
    "Recession",
    "Economic Growth",
    "Growth",
    "Housing market",
    "Job market",
    "Inflation",
    "Global economy",
    "EconDev",
    "Personal finance",
]

# identifies the script to Reddit — required or requests get blocked
USER_AGENT = "python:thesis-economic-discourse:1.0 (by /u/Cupcake-barry328625)"

NEWS_KEYWORDS      = {"news", "media", "journal", "times", "press", "worldnews"}
BUSINESS_KEYWORDS  = {"business", "corporate", "entrepreneur", "smallbusiness"}
DEDICATED_KEYWORDS = {"economy", "economic", "finance", "financial", "market",
                      "invest", "money", "wealth", "trading", "gdp", "econ",
                      "wallstreet", "stocks", "personalfinance"}


# ── helpers ───────────────────────────────────────────────────────────────────

def classify_category(subreddit: str) -> str:
    sub = subreddit.lower()
    if any(k in sub for k in NEWS_KEYWORDS):
        return "News outlet"
    if any(k in sub for k in BUSINESS_KEYWORDS):
        return "Business account"
    if any(k in sub for k in DEDICATED_KEYWORDS):
        return "Dedicated account"
    return "Personal account"


def is_english(text: str) -> bool:
    if not text or len(text.strip()) < 15:
        return False
    try:
        return detect(text) == "en"
    except LangDetectException:
        return False


def get_next_row_number(ws) -> int:
    for row in range(ws.max_row, DATA_START_ROW - 1, -1):
        val = ws.cell(row=row, column=1).value
        if val is not None:
            return int(val) + 1
    return 1


def get_already_scraped(ws) -> set:
    urls = set()
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if row[3]:
            urls.add(row[3])
    return urls


def search_reddit(client: httpx.Client, query: str, after: str = "") -> dict:
    params = {"q": query, "sort": "relevance", "limit": 25, "type": "link"}
    if after:
        params["after"] = after
    resp = client.get(
        "https://www.reddit.com/search.json",
        params=params,
        headers={"User-Agent": USER_AGENT},
    )
    resp.raise_for_status()
    return resp.json()


# ── main scrape ───────────────────────────────────────────────────────────────

def scrape(keywords: list, posts_per_keyword: int):
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[REDDIT_SHEET]

    row_number = get_next_row_number(ws)
    already_scraped = get_already_scraped(ws)
    today = date.today().strftime("%m/%d/%Y")

    total_keywords = len(keywords)
    session_start = time.time()
    keyword_times = []

    print(f"Starting at row #{row_number}. {len(already_scraped)} posts already in sheet.")
    print(f"Platform: Reddit | Keywords: {total_keywords} | Posts per keyword: {posts_per_keyword}\n")

    with httpx.Client(timeout=30) as client:
        for keyword_index, keyword in enumerate(keywords, start=1):
            pct = (keyword_index - 1) / total_keywords * 100
            print(f"\n[{keyword_index}/{total_keywords} — {pct:.0f}%] {keyword}")

            collected = 0
            skipped = 0
            after = ""
            keyword_start = time.time()

            try:
                while collected < posts_per_keyword:
                    data = search_reddit(client, keyword, after)
                    posts = data.get("data", {}).get("children", [])
                    after = data.get("data", {}).get("after") or ""

                    if not posts:
                        break

                    for post in posts:
                        if collected >= posts_per_keyword:
                            break

                        p = post.get("data", {})

                        # skip link/image posts — we need text content
                        if not p.get("is_self", False):
                            skipped += 1
                            continue

                        url = f"https://www.reddit.com{p.get('permalink', '')}"
                        if url in already_scraped:
                            continue

                        title = p.get("title", "").strip()
                        body = p.get("selftext", "").strip()

                        # skip deleted or removed posts
                        if body in ("[deleted]", "[removed]"):
                            body = ""

                        content = f"{title}\n\n{body}".strip() if body else title
                        if not content or not is_english(content):
                            skipped += 1
                            continue

                        subreddit = p.get("subreddit", "")
                        category = classify_category(subreddit)
                        likes = p.get("score", 0)
                        reposts = p.get("num_crossposts", 0)
                        comments = p.get("num_comments", 0)

                        created_ts = p.get("created_utc", 0)
                        date_posted = (
                            datetime.fromtimestamp(int(created_ts)).strftime("%m/%d/%Y")
                            if created_ts else ""
                        )

                        ws.append([
                            row_number,
                            keyword,
                            content,
                            url,
                            category,
                            likes,
                            reposts,
                            comments,
                            date_posted,
                            today,
                        ])

                        already_scraped.add(url)
                        row_number += 1
                        collected += 1
                        print(f"  [{collected}/{posts_per_keyword}] r/{subreddit} | {title[:65]}...")

                        time.sleep(max(0.5, random.gauss(1.2, 0.3)))

                    if not after:
                        print(f"  No more pages available for '{keyword}'.")
                        break

                    time.sleep(random.uniform(2, 5))

            except Exception as e:
                print(f"  Error on keyword '{keyword}': {e}")

            wb.save(XLSX_PATH)
            keyword_elapsed = time.time() - keyword_start
            keyword_times.append(keyword_elapsed)

            done_pct = keyword_index / total_keywords * 100
            avg_time = sum(keyword_times) / len(keyword_times)
            remaining = total_keywords - keyword_index
            eta_str = str(timedelta(seconds=int(avg_time * remaining)))
            elapsed_str = str(timedelta(seconds=int(time.time() - session_start)))

            print(f"  Saved. Collected {collected}, skipped {skipped} | "
                  f"keyword took {keyword_elapsed:.0f}s | "
                  f"overall {done_pct:.0f}% | elapsed {elapsed_str} | ETA ~{eta_str}")

            if remaining > 0:
                pause = random.uniform(3, 8)
                print(f"  Waiting {pause:.1f}s before next keyword...")
                time.sleep(pause)

    total_elapsed = str(timedelta(seconds=int(time.time() - session_start)))
    print(f"\nDone in {total_elapsed}. Results saved to {XLSX_PATH}")


def main():
    parser = argparse.ArgumentParser(
        prog="reddit_scraper",
        description="Scrape Reddit posts into the research xlsx.",
    )
    parser.add_argument(
        "-n", "--count",
        type=int, default=None, metavar="N",
        help=f"Posts per keyword (default: {POSTS_PER_KEYWORD})",
    )
    parser.add_argument(
        "-k", "--keywords",
        nargs="+", default=None, metavar="KEYWORD",
        help='Keywords to search (default: hardcoded 10)',
    )

    args = parser.parse_args()
    keywords = args.keywords or KEYWORDS
    posts_per_keyword = args.count or POSTS_PER_KEYWORD

    scrape(keywords, posts_per_keyword)


if __name__ == "__main__":
    main()
