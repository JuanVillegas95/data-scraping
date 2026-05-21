import argparse
import asyncio
import json
import random
import time
from datetime import date, datetime, timedelta
from pathlib import Path
from urllib.parse import unquote

import openpyxl
from langdetect import detect, LangDetectException
from playwright.async_api import async_playwright

from transcriber import transcribe_url

XLSX_PATH         = "Data gathering example(1).xlsx"
INSTAGRAM_SHEET   = "Instagram"
POSTS_PER_KEYWORD = 40
DATA_START_ROW    = 7   # rows 1-6 are account info + blanks + headers
COOKIES_FILE      = ".instagram_cookies.txt"

SESSION_ID = "15250725158%3AhV2zjCiS0g5VI4%3A16%3AAYjaw3RmNEdSvg1Ug4jhPZQTAAWmk0JvwZ2Q1zdtfA"
CSRF_TOKEN = "Zd9yOG3hdjbuHFTqQqdKoOGHJWJJgpBw"
DS_USER_ID = "15250725158"

KEYWORDS = {
    "Economy":          "economy",
    "Recession":        "recession",
    "Economic Growth":  "economicgrowth",
    "Growth":           "growth",
    "Housing market":   "housingmarket",
    "Job market":       "jobmarket",
    "Inflation":        "inflation",
    "Global economy":   "globaleconomy",
    "EconDev":          "econdev",
    "Personal finance": "personalfinance",
}

NEWS_KEYWORDS      = {"news", "media", "journal", "times", "press", "bbc", "cnn",
                      "reuters", "bloomberg", "economist", "daily", "herald", "report"}
BUSINESS_KEYWORDS  = {"official", "corp", "inc", "ltd", "company", "bank",
                      "invest", "capital", "fund", "group"}
DEDICATED_KEYWORDS = {"economy", "economic", "finance", "financial", "market",
                      "stock", "money", "wealth", "trading", "crypto", "gdp",
                      "econ", "biz", "business"}


# ── helpers ───────────────────────────────────────────────────────────────────

def write_cookies_file():
    """Write a Netscape-format cookies file so yt-dlp can authenticate."""
    expire = "1874822400"  # 2029-01-01 — well past session expiry
    sid    = unquote(SESSION_ID)
    lines  = [
        "# Netscape HTTP Cookie File",
        f".instagram.com\tTRUE\t/\tTRUE\t{expire}\tsessionid\t{sid}",
        f".instagram.com\tTRUE\t/\tTRUE\t{expire}\tcsrftoken\t{CSRF_TOKEN}",
        f".instagram.com\tTRUE\t/\tFALSE\t{expire}\tds_user_id\t{DS_USER_ID}",
    ]
    Path(COOKIES_FILE).write_text("\n".join(lines))


def classify_category(username: str, verified: bool = False) -> str:
    text = username.lower()
    if any(k in text for k in NEWS_KEYWORDS):
        return "News outlet"
    if verified or any(k in text for k in BUSINESS_KEYWORDS):
        return "Business account"
    if any(k in text for k in DEDICATED_KEYWORDS):
        return "Dedicated account"
    return "Personal account"


def is_english(text: str) -> bool:
    if not text or len(text.strip()) < 15:
        return False
    try:
        return detect(text) == "en"
    except LangDetectException:
        return False


def get_next_row_number(ws) -> int:
    for row in range(ws.max_row, DATA_START_ROW - 1, -1):
        val = ws.cell(row=row, column=1).value
        if val is not None:
            return int(val) + 1
    return 1


def get_already_scraped(ws) -> set:
    urls = set()
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if row[3]:
            urls.add(row[3])
    return urls


async def get_post_metadata(url: str) -> dict | None:
    """Run yt-dlp --dump-json to get post metadata without downloading."""
    cmd = [
        "yt-dlp", "--dump-json", "--no-download",
        "--cookies", COOKIES_FILE,
        "--quiet", "--no-warnings",
        "--socket-timeout", "20",
        url,
    ]
    proc = await asyncio.create_subprocess_exec(
        *cmd,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
    )
    try:
        stdout, _ = await asyncio.wait_for(proc.communicate(), timeout=30)
    except asyncio.TimeoutError:
        proc.kill()
        return None

    if proc.returncode != 0:
        return None
    try:
        return json.loads(stdout)
    except Exception:
        return None


# ── playwright link collector ─────────────────────────────────────────────────

async def collect_post_links(hashtag: str, need: int) -> list[str]:
    """Open Instagram hashtag page and collect /p/ links by scrolling."""
    links: set[str] = set()

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (X11; Linux x86_64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            )
        )
        await context.add_cookies([
            {"name": "sessionid",  "value": unquote(SESSION_ID), "domain": ".instagram.com", "path": "/"},
            {"name": "csrftoken",  "value": CSRF_TOKEN,           "domain": ".instagram.com", "path": "/"},
            {"name": "ds_user_id", "value": DS_USER_ID,           "domain": ".instagram.com", "path": "/"},
        ])

        page = await context.new_page()
        await page.goto(
            f"https://www.instagram.com/explore/tags/{hashtag}/",
            wait_until="networkidle",
            timeout=60000,
        )
        await asyncio.sleep(6)

        # debug: report where we actually ended up
        actual_url = page.url
        all_links  = await page.query_selector_all("a[href]")
        print(f"  [debug] landed on: {actual_url}")
        print(f"  [debug] total <a> tags on page: {len(all_links)}")

        for _ in range(25):
            if len(links) >= need * 3:
                break

            # try both /p/ and /reel/ links
            for sel in ('a[href*="/p/"]', 'a[href*="/reel/"]'):
                els = await page.query_selector_all(sel)
                for el in els:
                    href = await el.get_attribute("href")
                    if href:
                        url = ("https://www.instagram.com" + href
                               if href.startswith("/") else href)
                        url = url.split("?")[0].rstrip("/") + "/"
                        links.add(url)

            await page.evaluate("window.scrollBy(0, window.innerHeight * 3)")
            await asyncio.sleep(random.uniform(2, 4))

        await browser.close()

    print(f"  Collected {len(links)} post links from hashtag page.")
    return list(links)


# ── main scrape ───────────────────────────────────────────────────────────────

async def scrape(keywords: dict, posts_per_keyword: int):
    write_cookies_file()

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[INSTAGRAM_SHEET]

    row_number      = get_next_row_number(ws)
    already_scraped = get_already_scraped(ws)
    today           = date.today().strftime("%m/%d/%Y")

    total_keywords = len(keywords)
    session_start  = time.time()
    keyword_times  = []

    print(f"Starting at row #{row_number}. {len(already_scraped)} posts already in sheet.")
    print(f"Platform: Instagram | Keywords: {total_keywords} | Posts per keyword: {posts_per_keyword}\n")

    for keyword_index, (keyword, hashtag) in enumerate(keywords.items(), start=1):
        pct = (keyword_index - 1) / total_keywords * 100
        print(f"\n[{keyword_index}/{total_keywords} — {pct:.0f}%] {keyword} (#{hashtag})")

        collected     = 0
        skipped       = 0
        keyword_start = time.time()

        try:
            post_links = await collect_post_links(hashtag, posts_per_keyword)

            for url in post_links:
                if collected >= posts_per_keyword:
                    break
                if url in already_scraped:
                    continue

                # ── get post metadata via yt-dlp ─────────────────────────
                meta = await get_post_metadata(url)
                if not meta:
                    skipped += 1
                    continue

                caption  = (meta.get("description") or meta.get("title") or "").strip()
                vcodec   = meta.get("vcodec", "none")
                is_video = vcodec not in (None, "none", "")

                username = meta.get("uploader_id") or meta.get("uploader") or ""
                verified = meta.get("channel_is_verified", False)
                likes    = meta.get("like_count") or 0
                comments = meta.get("comment_count") or 0
                ts       = meta.get("timestamp") or 0
                date_posted = (
                    datetime.fromtimestamp(int(ts)).strftime("%m/%d/%Y") if ts else ""
                )

                # ── get content ──────────────────────────────────────────
                content = None
                source  = None

                if is_video:
                    print(f"    video, transcribing...")
                    content = await transcribe_url(url, cookies_file=COOKIES_FILE)
                    if content:
                        source = "whisper"
                    elif caption and is_english(caption):
                        content = caption
                        source  = "caption-fallback"
                else:
                    if caption and is_english(caption):
                        content = caption
                        source  = "caption"

                if not content:
                    skipped += 1
                    continue

                category = classify_category(username, verified=verified)

                ws.append([
                    row_number,
                    keyword,
                    content,
                    url,
                    category,
                    likes,
                    0,   # Instagram has no repost count
                    comments,
                    date_posted,
                    today,
                ])

                already_scraped.add(url)
                row_number += 1
                collected  += 1
                print(f"  [{collected}/{posts_per_keyword}] ({source}) "
                      f"@{username} | {content[:65]}...")

                await asyncio.sleep(max(1.5, random.gauss(3.0, 1.0)))

        except Exception as e:
            print(f"  Error on keyword '{keyword}': {e}")

        wb.save(XLSX_PATH)
        keyword_elapsed = time.time() - keyword_start
        keyword_times.append(keyword_elapsed)

        done_pct  = keyword_index / total_keywords * 100
        avg_time  = sum(keyword_times) / len(keyword_times)
        remaining = total_keywords - keyword_index
        eta_str   = str(timedelta(seconds=int(avg_time * remaining)))
        elapsed   = str(timedelta(seconds=int(time.time() - session_start)))

        print(f"  Saved. Collected {collected}, skipped {skipped} | "
              f"keyword took {keyword_elapsed:.0f}s | "
              f"overall {done_pct:.0f}% | elapsed {elapsed} | ETA ~{eta_str}")

        if remaining > 0:
            pause = random.uniform(10, 20)
            print(f"  Waiting {pause:.1f}s before next keyword...")
            await asyncio.sleep(pause)

    # clean up cookies file
    Path(COOKIES_FILE).unlink(missing_ok=True)

    total_elapsed = str(timedelta(seconds=int(time.time() - session_start)))
    print(f"\nDone in {total_elapsed}. Results saved to {XLSX_PATH}")


def main():
    parser = argparse.ArgumentParser(
        prog="instagram_scraper",
        description="Scrape Instagram posts into the research xlsx.",
    )
    parser.add_argument("-n", "--count", type=int, default=None, metavar="N",
                        help=f"Posts per keyword (default: {POSTS_PER_KEYWORD})")
    parser.add_argument("-k", "--keywords", nargs="+", default=None, metavar="KEYWORD",
                        help="Keywords to search (default: hardcoded 10)")

    args  = parser.parse_args()
    kws   = ({kw: kw.lower().replace(" ", "") for kw in args.keywords}
             if args.keywords else KEYWORDS)
    count = args.count or POSTS_PER_KEYWORD

    asyncio.run(scrape(kws, count))


if __name__ == "__main__":
    main()
