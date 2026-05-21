import argparse
import asyncio
import os
import random
import re
import time
import urllib.request
from datetime import date, datetime, timedelta

import openpyxl
from langdetect import detect, LangDetectException
from TikTokApi import TikTokApi
from transcriber import transcribe_url

XLSX_PATH = "Data gathering example(1).xlsx"
TIKTOK_SHEET = "Tiktok"
POSTS_PER_KEYWORD = 40
DATA_START_ROW = 5  # rows 1-4 are account info + headers

# keyword label → hashtag name
KEYWORDS = {
    "Economy": "economy",
    "Recession": "recession",
    "Economic Growth": "economicgrowth",
    "Growth": "growth",
    "Housing market": "housingmarket",
    "Job market": "jobmarket",
    "Inflation": "inflation",
    "Global economy": "globaleconomy",
    "EconDev": "econdev",
    "Personal finance": "personalfinance",
}

NEWS_KEYWORDS = {
    "news", "media", "press", "journal", "times", "post", "telegraph",
    "guardian", "bbc", "cnn", "reuters", "bloomberg", "economist",
    "wsj", "nyt", "today", "daily", "herald", "report", "channel",
}
BUSINESS_KEYWORDS = {
    "official", "corp", "inc", "ltd", "company", "bank",
    "invest", "capital", "fund", "asset", "group",
}
DEDICATED_KEYWORDS = {
    "economy", "economic", "finance", "financial", "market",
    "stock", "money", "wealth", "trading", "crypto", "fiscal",
    "monetary", "recession", "gdp", "econ", "biz", "business",
}

MS_TOKEN = "BxCMhtl79mFqfzDm-WhEfrrZYqB-4vanronP311JvRTeCh-szsQeGkUn-Qz9WmNJ4pwYl7YK0RzGJpVazN9Cd_ylFLZL1YMh_jiY6h4kSuAvZhsacuyTlhBEVGbSNTdFecH2S2p1fr4="  # expires 2026-05-30


# ── helpers ──────────────────────────────────────────────────────────────────

def classify_category(author: dict) -> str:
    nickname = (author.get("nickname") or "").lower()
    bio = (author.get("signature") or "").lower()
    username = (author.get("uniqueId") or "").lower()
    verified = author.get("verified", False)
    text = f"{nickname} {bio} {username}"

    if any(k in text for k in NEWS_KEYWORDS):
        return "News outlet"
    if verified or any(k in text for k in BUSINESS_KEYWORDS):
        return "Business account"
    if any(k in text for k in DEDICATED_KEYWORDS):
        return "Dedicated account"
    return "Personal account"


def is_english(text: str) -> bool:
    if not text or len(text.strip()) < 15:
        return False
    try:
        return detect(text) == "en"
    except LangDetectException:
        return False


def parse_webvtt(content: str) -> str:
    """Extract clean plain text from a WebVTT subtitle file."""
    lines = []
    for line in content.splitlines():
        line = line.strip()
        if not line or line.startswith("WEBVTT") or line.startswith("NOTE"):
            continue
        if re.match(r"^\d+$", line) or "-->" in line:
            continue
        line = re.sub(r"<[^>]+>", "", line)  # strip inline tags like <c>, <00:00:01>
        if line:
            lines.append(line)

    # deduplicate consecutive identical lines (overlapping captions)
    deduped = []
    for line in lines:
        if not deduped or line != deduped[-1]:
            deduped.append(line)
    return " ".join(deduped)


async def fetch_subtitles(url: str) -> str | None:
    def _fetch():
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=10) as r:
                return r.read().decode("utf-8")
        except Exception:
            return None

    content = await asyncio.to_thread(_fetch)
    if not content:
        return None
    return parse_webvtt(content) or None


def find_english_subtitle_url(video_data: dict) -> str | None:
    """Return the URL of the English subtitle track if TikTok has one."""
    for sub in video_data.get("subtitleInfos", []):
        lang = (sub.get("LanguageCodeName") or sub.get("languageCodeName") or "").lower()
        if lang.startswith("en"):
            return sub.get("Url") or sub.get("url")
    return None


# ── sheet helpers ─────────────────────────────────────────────────────────────

def get_next_row_number(ws) -> int:
    for row in range(ws.max_row, DATA_START_ROW - 1, -1):
        val = ws.cell(row=row, column=1).value
        if val is not None:
            return int(val) + 1
    return 1


def get_already_scraped(ws) -> set:
    urls = set()
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if row[3]:
            urls.add(row[3])
    return urls


# ── main scrape ───────────────────────────────────────────────────────────────

async def scrape(ms_token: str | None, keywords: dict = None, posts_per_keyword: int = None):
    keywords = keywords or KEYWORDS
    posts_per_keyword = posts_per_keyword or POSTS_PER_KEYWORD

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[TIKTOK_SHEET]

    row_number = get_next_row_number(ws)
    already_scraped = get_already_scraped(ws)
    today = date.today().strftime("%m/%d/%Y")

    total_keywords = len(keywords)
    session_start = time.time()
    keyword_times = []

    print(f"Starting at row #{row_number}. {len(already_scraped)} posts already in sheet.")
    print(f"Platform: TikTok | Keywords: {total_keywords} | Posts per keyword: {posts_per_keyword}\n")

    async with TikTokApi() as api:
        await api.create_sessions(
            ms_tokens=[ms_token] if ms_token else None,
            num_sessions=1,
            sleep_after=3,
            headless=False,
        )

        for keyword_index, (keyword, hashtag) in enumerate(keywords.items(), start=1):
            pct = (keyword_index - 1) / total_keywords * 100
            print(f"\n[{keyword_index}/{total_keywords} — {pct:.0f}%] {keyword} (#{hashtag})")
            collected = 0
            skipped = 0
            keyword_start = time.time()

            try:
                # fetch 3x more than needed to account for non-English skips
                async for video in api.hashtag(name=hashtag).videos(count=posts_per_keyword * 3):
                    if collected >= posts_per_keyword:
                        break

                    d = video.as_dict
                    author = d.get("author", {})
                    stats = d.get("statsV2") or d.get("stats") or {}

                    video_id = d.get("id", "")
                    username = author.get("uniqueId", "")
                    url = f"https://www.tiktok.com/@{username}/video/{video_id}"

                    if url in already_scraped:
                        continue

                    # ── get content ────────────────────────────────────────
                    # 1. TikTok auto-captions (fastest, free)
                    # 2. yt-dlp + Whisper (fallback, slower)
                    # 3. skip — no description fallback
                    content = None
                    source = None

                    sub_url = find_english_subtitle_url(d)
                    if sub_url:
                        transcription = await fetch_subtitles(sub_url)
                        if transcription and is_english(transcription):
                            content = transcription
                            source = "tiktok-captions"

                    if not content:
                        print(f"    no captions, trying yt-dlp + Whisper...")
                        content = await transcribe_url(url)
                        if content:
                            source = "whisper"

                    if not content:
                        skipped += 1
                        continue

                    likes = stats.get("diggCount", 0)
                    reposts = stats.get("shareCount", 0)
                    comments = stats.get("commentCount", 0)

                    create_ts = d.get("createTime", 0)
                    date_posted = (
                        datetime.fromtimestamp(int(create_ts)).strftime("%m/%d/%Y")
                        if create_ts else ""
                    )

                    category = classify_category(author)

                    ws.append([
                        row_number,
                        keyword,
                        content,
                        url,
                        category,
                        likes,
                        reposts,
                        comments,
                        date_posted,
                        today,
                    ])

                    already_scraped.add(url)
                    row_number += 1
                    collected += 1
                    print(f"  [{collected}/{posts_per_keyword}] ({source}) @{username} | {content[:65]}...")

                    await asyncio.sleep(max(1.5, random.gauss(3.0, 1.0)))

            except Exception as e:
                print(f"  Error on keyword '{keyword}': {e}")

            wb.save(XLSX_PATH)
            keyword_elapsed = time.time() - keyword_start
            keyword_times.append(keyword_elapsed)

            done_pct = keyword_index / total_keywords * 100
            avg_time = sum(keyword_times) / len(keyword_times)
            remaining = total_keywords - keyword_index
            eta_secs = avg_time * remaining
            eta_str = str(timedelta(seconds=int(eta_secs)))
            elapsed_str = str(timedelta(seconds=int(time.time() - session_start)))

            print(f"  Saved. Collected {collected}, skipped {skipped} | "
                  f"keyword took {keyword_elapsed:.0f}s | "
                  f"overall {done_pct:.0f}% | elapsed {elapsed_str} | ETA ~{eta_str}")

            if remaining > 0:
                pause = random.uniform(10, 25)
                print(f"  Waiting {pause:.1f}s before next keyword...")
                await asyncio.sleep(pause)

    total_elapsed = str(timedelta(seconds=int(time.time() - session_start)))
    print(f"\nDone in {total_elapsed}. Results saved to {XLSX_PATH}")


def main():
    parser = argparse.ArgumentParser(
        prog="tiktok_scraper",
        description="Scrape social media posts into the research xlsx.",
    )
    parser.add_argument(
        "-p", "--platform",
        choices=["tiktok"],
        default="tiktok",
        help="Platform to scrape (default: tiktok)",
    )
    parser.add_argument(
        "-n", "--count",
        type=int,
        default=None,
        metavar="N",
        help=f"Posts per keyword (default: {POSTS_PER_KEYWORD})",
    )
    parser.add_argument(
        "-k", "--keywords",
        nargs="+",
        default=None,
        metavar="KEYWORD",
        help='Keywords to search, e.g. -k "economy" "inflation" (default: hardcoded 10)',
    )

    args = parser.parse_args()

    # build keyword dict: label → hashtag (lowercase, no spaces)
    if args.keywords:
        keywords = {kw: kw.lower().replace(" ", "") for kw in args.keywords}
    else:
        keywords = None  # scrape() will use the hardcoded KEYWORDS default

    ms_token = os.getenv("TIKTOK_MS_TOKEN") or MS_TOKEN
    asyncio.run(scrape(ms_token, keywords=keywords, posts_per_keyword=args.count))


if __name__ == "__main__":
    main()
